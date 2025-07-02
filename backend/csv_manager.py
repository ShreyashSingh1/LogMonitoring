import pandas as pd
import os
import json
from datetime import datetime
import threading
import sqlite3
from openpyxl import load_workbook

# --- LOG EXPORT CONFIGURATION ---
# Set these to True/False to enable/disable log export formats
EXPORT_JSONL = True
EXPORT_XLSX = True
# --------------------------------

class CSVManager:
    def __init__(self):
        self.csv_dir = os.path.join(os.path.dirname(__file__), 'csv_data')
        self.db_path = os.path.join(self.csv_dir, 'logs.db')
        self.lock = threading.Lock()
        
        # Create directories
        os.makedirs(self.csv_dir, exist_ok=True)
        
        # Initialize database
        self.init_database()
    
    def init_database(self):
        """Initialize SQLite database for fast querying"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute('''
                CREATE TABLE IF NOT EXISTS logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source TEXT,
                    log_type TEXT,
                    level TEXT,
                    message TEXT,
                    timestamp TEXT,
                    file_path TEXT,
                    parsed_at TEXT,
                    raw_content TEXT,
                    user_id TEXT,
                    method TEXT,
                    url TEXT,
                    status_code INTEGER,
                    client_ip TEXT,
                    duration_ms REAL,
                    request_id TEXT,
                    exception_type TEXT,
                    exception_message TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Create indexes for faster queries
            conn.execute('CREATE INDEX IF NOT EXISTS idx_level ON logs(level)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_source ON logs(source)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_timestamp ON logs(timestamp)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_created_at ON logs(created_at)')
    
    def _prepare_for_csv(self, log_dict):
        """Convert all values to strings (except None), serialize dicts/lists to JSON."""
        result = {}
        for key, value in log_dict.items():
            if isinstance(value, dict) or isinstance(value, list):
                result[key] = json.dumps(value, ensure_ascii=False)
            elif value is None:
                result[key] = ''
            elif not isinstance(value, str):
                result[key] = str(value)
            else:
                result[key] = value
        return result
    
    def add_log(self, parsed_log):
        """Add a parsed log entry to CSV and database"""
        try:
            with self.lock:
                # Serialize dict fields to JSON strings
                parsed_log = self._prepare_for_csv(parsed_log)
                # Add to database
                self.add_to_database(parsed_log)
                # Add to CSV files
                self.add_to_csv(parsed_log)
        except Exception as e:
            print(f"Error adding log to CSV: {e}")
    
    def add_to_database(self, parsed_log):
        """Add log to SQLite database"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute('''
                INSERT INTO logs (
                    source, log_type, level, message, timestamp, file_path, 
                    parsed_at, raw_content, user_id, method, url, status_code, 
                    client_ip, duration_ms, request_id, exception_type, exception_message
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                parsed_log.get('source'),
                parsed_log.get('log_type'),
                parsed_log.get('level'),
                parsed_log.get('message'),
                parsed_log.get('timestamp'),
                parsed_log.get('file_path'),
                parsed_log.get('parsed_at'),
                parsed_log.get('raw_content'),
                parsed_log.get('user_id'),
                parsed_log.get('method'),
                parsed_log.get('url'),
                parsed_log.get('status_code'),
                parsed_log.get('client_ip'),
                parsed_log.get('duration_ms'),
                parsed_log.get('request_id'),
                parsed_log.get('exception_type'),
                parsed_log.get('exception_message')
            ))
    
    def add_to_csv(self, parsed_log):
        """Add log to appropriate JSONL and XLSX files (CSV writing disabled)"""
        csv_ready_log = self._prepare_for_csv(parsed_log)
        # General logs JSONL
        if EXPORT_JSONL:
            general_jsonl = os.path.join(self.csv_dir, 'all_logs.jsonl')
            self.append_to_jsonl(general_jsonl, parsed_log)
        # General logs XLSX
        if EXPORT_XLSX:
            general_xlsx = os.path.join(self.csv_dir, 'all_logs.xlsx')
            self.append_to_xlsx(general_xlsx, parsed_log)
        # Source-specific JSONL
        if EXPORT_JSONL:
            source_jsonl = os.path.join(self.csv_dir, f'{csv_ready_log["source"]}_logs.jsonl')
            self.append_to_jsonl(source_jsonl, parsed_log)
        # Source-specific XLSX
        if EXPORT_XLSX:
            source_xlsx = os.path.join(self.csv_dir, f'{csv_ready_log["source"]}_logs.xlsx')
            self.append_to_xlsx(source_xlsx, parsed_log)
        # Error logs JSONL/XLSX
        if csv_ready_log.get('level', '').lower() in ['error', 'warn', 'warning', 'fatal']:
            if EXPORT_JSONL:
                error_jsonl = os.path.join(self.csv_dir, 'error_logs.jsonl')
                self.append_to_jsonl(error_jsonl, parsed_log)
            if EXPORT_XLSX:
                error_xlsx = os.path.join(self.csv_dir, 'error_logs.xlsx')
                self.append_to_xlsx(error_xlsx, parsed_log)
        # Daily JSONL
        if EXPORT_JSONL:
            date_str = datetime.now().strftime('%Y-%m-%d')
            daily_jsonl = os.path.join(self.csv_dir, f'logs_{date_str}.jsonl')
            self.append_to_jsonl(daily_jsonl, parsed_log)
        # Daily XLSX
        if EXPORT_XLSX:
            date_str = datetime.now().strftime('%Y-%m-%d')
            daily_xlsx = os.path.join(self.csv_dir, f'logs_{date_str}.xlsx')
            self.append_to_xlsx(daily_xlsx, parsed_log)
    
    def append_to_jsonl(self, jsonl_path, parsed_log):
        """Append a single log entry to a JSON Lines file."""
        try:
            with open(jsonl_path, 'a', encoding='utf-8') as f:
                json.dump(parsed_log, f, ensure_ascii=False)
                f.write('\n')
        except Exception as e:
            print(f"Error appending to JSONL {jsonl_path}: {e}")
    
    def append_to_xlsx(self, xlsx_path, parsed_log):
        """Append a single log entry to an Excel file (.xlsx) using pandas."""
        try:
            df = pd.DataFrame([self._prepare_for_csv(parsed_log)])
            if not os.path.exists(xlsx_path):
                # Create new file with header
                df.to_excel(xlsx_path, index=False)
            else:
                # Append to existing file
                book = load_workbook(xlsx_path)
                writer = pd.ExcelWriter(xlsx_path, engine='openpyxl')
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                # Find the last row in the existing sheet
                if 'Sheet1' in writer.sheets:
                    startrow = writer.sheets['Sheet1'].max_row
                else:
                    # If Sheet1 does not exist, create it
                    startrow = 0
                df.to_excel(writer, index=False, header=False, startrow=startrow)
                writer.save()
                writer.close()
        except Exception as e:
            print(f"Error appending to XLSX {xlsx_path}: {e}")
    
    def get_logs(self, log_type='all', level='all', limit=100, source=None):
        """Get logs from database with filtering"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                
                query = "SELECT * FROM logs WHERE 1=1"
                params = []
                
                if log_type != 'all':
                    query += " AND log_type = ?"
                    params.append(log_type)
                
                if level != 'all':
                    query += " AND level = ?"
                    params.append(level)
                
                if source:
                    query += " AND source = ?"
                    params.append(source)
                
                query += " ORDER BY created_at DESC LIMIT ?"
                params.append(limit)
                
                cursor = conn.execute(query, params)
                rows = cursor.fetchall()
                
                return [dict(row) for row in rows]
                
        except Exception as e:
            print(f"Error getting logs: {e}")
            return []
    
    def get_error_logs(self, limit=100):
        """Get only error and warning logs"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                
                cursor = conn.execute('''
                    SELECT * FROM logs 
                    WHERE level IN ('error', 'warn', 'warning', 'fatal')
                    ORDER BY created_at DESC 
                    LIMIT ?
                ''', (limit,))
                
                rows = cursor.fetchall()
                return [dict(row) for row in rows]
                
        except Exception as e:
            print(f"Error getting error logs: {e}")
            return []
    
    def get_statistics(self):
        """Get log statistics"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                # Total logs
                total_logs = conn.execute("SELECT COUNT(*) FROM logs").fetchone()[0]
                # Logs by level
                level_stats = {}
                cursor = conn.execute("SELECT level, COUNT(*) FROM logs GROUP BY level")
                for row in cursor:
                    level_stats[row[0]] = row[1]
                # Logs by source
                source_stats = {}
                cursor = conn.execute("SELECT source, COUNT(*) FROM logs GROUP BY source")
                for row in cursor:
                    source_stats[row[0]] = row[1]
                # Recent activity (last hour)
                cursor = conn.execute('''
                    SELECT COUNT(*) FROM logs 
                    WHERE created_at > datetime('now', '-1 hour')
                ''')
                recent_logs = cursor.fetchone()[0]
                return {
                    'total_logs': total_logs,
                    'level_distribution': level_stats,
                    'source_distribution': source_stats,
                    'recent_activity': recent_logs
                }
        except Exception as e:
            print(f"Error getting statistics: {e}")
            return {} 